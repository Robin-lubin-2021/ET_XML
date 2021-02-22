import xml.etree.ElementTree as ET
#输入模块os
import os
import re
import openpyxl
xml_path = "L:C:\\Users\\LUBIN\\Desktop\\Data Portal Test"
xml_name_sum = os.listdir(xml_path)
#print(xml_name_summ)
manu_tree = ET.parse("{}\MANUFACTORS.xml".format(xml_path))  
manu_root  = manu_tree.getroot()
#print(manu_root.tag, manu_root.attrib)
xlsx_path = "C:\\Users\\LUBIN\\Desktop\\xlsx_test"
xlsx_name_sum = os.listdir(xlsx_path)
for xml_name in xml_name_sum:
        re_com0 = re.compile("\.xml$")
        xml_name_search = re_com0.search(xml_name)
        #print(xml_name_search)
        if  xml_name_search is not None:
            tree = ET.parse("{}\\{}".format(xml_path, xml_name))
            root = tree.getroot()
            #print(root.tag, root.attrib)
            #root.set("count", "1")
            #root.set("build", "136.22")
            for child in root:
                #print(child.tag)
                part_number_xml = child.get("P_ARTICLE_PARTNR")
                if part_number_xml is not None:
                    #print(part_number_xml)
                    re_com1 = re.compile("^[\w]+[-]*[\w]+\.")
                    type_shortname_match = re_com1.match(part_number_xml)
                    type_shortname = type_shortname_match.group()
                    #print(type_shortname)
                    manu_shortname1 = type_shortname[:len(type_shortname)-1]
                    #print(manu_shortname1, end = ' ')
                    for manu_child in manu_root:
                            manu_shortname2 = manu_child.get("P_PART_ADDRESS_SHORTNAME") 
                            if manu_shortname2 == manu_shortname1:
                                #print(manu_shortname2, end = ' ')
                                manu_longname = manu_child.get("P_PART_ADDRESS_LONGNAME")
                                #print(manu_longname)
                                wb = openpyxl.load_workbook("{}\\{}.xlsx".format(xlsx_path, manu_longname))
                                type_number_xml = child.get("P_ARTICLE_TYPENR")
                                #print(type_number_xml)
                                sh = wb["Sheet1"]
                                for i in range(sh.min_row, sh.max_row):
                                    type_value_xlsx = sh.cell(row = i, column = 4).value
                                    if type_value_xlsx == type_number_xml:
                                        print(type_number_xml)
                                        price_value_xlsx = round(sh.cell(row = i, column = 8).value, 2)
                                        #print((price_value_xlsx))
                                        child.set("P_ARTICLE_PURCHASEPRICE_1", str(price_value_xlsx))
                                        part_descr1_xml = child.get("P_ARTICLE_DESCR1")
                                        if part_descr1_xml is not None:
                                            #print(part_descr1_xml)
                                            re_com2 = re.compile("(已询未税价)")
                                            if re_com2.search(part_descr1_xml) is not None:
                                                #print(re_com2.search(part_descr1_xml).group())
                                                part_descr1_xml = re_com2.sub("inquired price without tax", part_descr1_xml)
                                            """
                                            re_com3 = re.compile("es_ES@")
                                            if re_com3.search(part_descr1_xml) is not None:
                                                #print(re_com3.search(part_descr1_xml).group())
                                                part_descr1_xml = re_com3.sub("??_??@", part_descr1_xml)
                                            """
                                        child.set("P_ARTICLE_DESCR1", part_descr1_xml)  
                                        break
                                else:
                                    print("{} is not in {}.xlsx".format(type_number_xml, manu_longname))
                                wb.save(r"{}\\{}.xlsx".format(xlsx_path, manu_longname))
                                break
                    else:
                        print("{} is not in MANUFACTORS.xml".format(manu_shortname1))
            tree.write("{}\\{}".format(xml_path, xml_name))
manu_tree.write("{}\MANUFACTORS.xml".format(xml_path))
